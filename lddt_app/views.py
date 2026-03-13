from datetime import datetime
from django.db.models import Q
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .forms import *
from .filters import WebsiteFilter, VmFilter, shortwebsiteFilter, shortvmFilter
from .models import *
from datetime import *
from django.db.models import Max, Min
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import base64
from datetime import date, timedelta, datetime
from datetime import date, timedelta, datetime
import calendar
from django.core.cache import cache
from django.utils import timezone  # ✅  correct import
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.core.management import call_command
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from django.http import JsonResponse
from collections import defaultdict
import json
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import RunReportRequest, RunRealtimeReportRequest, DateRange, Metric
from google.oauth2 import service_account
import os
from google.analytics.data_v1beta import (
    BetaAnalyticsDataClient,
    RunReportRequest,
    RunRealtimeReportRequest,
    DateRange,
    Metric,
    Dimension,  # ✅  ADDED
)

from django.core.cache import cache
from django.shortcuts import render
import json
from dateutil.relativedelta import relativedelta

from django.db.models import OuterRef, Subquery
from django.shortcuts import render
from django.utils import timezone

from .models import GoogleAnalyticsStats

import logging
logger = logging.getLogger(__name__)

@login_required
#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    HOME
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################
def home(request):
    # --------------websites
    websites = Website.objects.filter(
        Q(type__name__exact="SITE") & ~Q(activity__name__exact="DECOMMISSIONED")
    ).order_by('ssl_expiry_date')
    count_all_sites = Website.objects.filter(
        Q(type__name__exact="SITE")
    ).count
    count_all_active = Website.objects.filter(
        Q(type__name__exact="SITE") & Q(activity__name__exact="ACTIVE")
    ).count
    count_sites_workingon = Website.objects.filter(
        Q(type__name__exact="SITE") & Q(tech_status__name__exact="WORKING ON")
    ).count
    count_sites_cert_sent = Website.objects.filter(
        Q(type__name__exact="SITE") & Q(tech_status__name__exact="CERT SENT")
    ).count

    count_sites_issue = Website.objects.filter(
        Q(type__name__exact="SITE") & Q(tech_status__name__exact="ISSUE")
    ).count
    # -----------------subsites
    count_all_subsites = Website.objects.filter(Q(type__name__exact="SUB-SITE")).count
    count_all_active_subsites = Website.objects.filter(
        Q(type__name__exact="SUB-SITE") & Q(activity__name__exact="ACTIVE")
    ).count
    count_subsites_workingon = Website.objects.filter(
        Q(type__name__exact="SUB-SITE") & Q(tech_status__name__exact="WORKING ON")
    ).count
    count_subsites_issue = Website.objects.filter(
        Q(type__name__exact="SUB-SITE") & Q(tech_status__name__exact="ISSUE")
    ).count
    # -----------------vm's
    table_item_count_vms = Vm.objects.all().count
    vms_working_on_status_count = Vm.objects.filter(
        Q(vm_type__name__exact="VM") & Q(vm_status__name__exact="Working On")
    ).count

    current_datetime_now = datetime.now()
    current_datetime_now_tostring = current_datetime_now

    return render(request, 'index.html', {
        'count_all_sites': count_all_sites,
        'count_all_active': count_all_active,
        'count_sites_workingon': count_sites_workingon,
        'count_sites_issue': count_sites_issue,
        'websites': websites,
        # subsites
        'count_all_subsites': count_all_subsites,
        'count_all_active_subsites': count_all_active_subsites,
        'count_subsites_workingon': count_subsites_workingon,
        'count_subsites_issue': count_subsites_issue,
        'current_datetime_now_tostring': current_datetime_now_tostring,
        # vm's
        'table_item_count_vms': table_item_count_vms,
        'count_sites_cert_sent': count_sites_cert_sent,
        'vms_working_on_status_count': vms_working_on_status_count,

    })

#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    WEBSITES
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def websites_home(request):
    websites = Website.objects.filter(
        Q(type__name__exact="SITE") & ~Q(activity__name__exact="DECOMMISSIONED")
    ).order_by('ssl_expiry_date')
    table_item_count_sites = Website.objects.filter(
        Q(type__name__exact="SITE") & ~Q(activity__name__exact="DECOMMISSIONED")
    ).count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs
    current_datetime_now = datetime.now()
    current_datetime_now_tostring = current_datetime_now.strftime("%Y/%B")

    return render(request, 'website/websites_home.html', {
        'websites': websites,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_sites': table_item_count_sites,
        'current_datetime_now_tostring': current_datetime_now_tostring,
    })


def websites_ssl_exp_this_month(request):
    websites = Website.objects.filter(
        Q(type__name__exact="SITE")
    ).order_by('ssl_expiry_date')
    table_item_count_sites = Website.objects.filter(
        Q(type__name__exact="SITE")
    ).count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs
    current_datetime_now = datetime.now()
    current_datetime_now_tostring = current_datetime_now.strftime("%Y/%B")

    return render(request, 'website/websites_expiry_this_month.html', {
        'websites': websites,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_sites': table_item_count_sites,
        'current_datetime_now_tostring': current_datetime_now_tostring,
    })


def websites_investigate_decommissioned(request):
    websites = Website.objects.filter(
        Q(type__name__exact="SITE") & Q(activity__name__exact="INVESTIGATE") | Q(activity__name__exact="DECOMMISSIONED")
    ).order_by('ssl_expiry_date')
    table_item_count_sites = Website.objects.filter(
        Q(type__name__exact="SITE")
    ).count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs
    current_datetime_now = datetime.now()
    current_datetime_now_tostring = current_datetime_now.strftime("%Y/%B")
    return render(request, 'website/websites_investigate_decommissioned.html', {
        'websites': websites,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_sites': table_item_count_sites,
        'current_datetime_now_tostring': current_datetime_now_tostring,
    })


def websites_google_analytics(request):
    websites = Website.objects.filter(
        ~Q(activity__name__exact="DECOMMISSIONED") & Q(ga4_required__name__exact="Yes")
    ).order_by('ssl_expiry_date')
    table_item_count = Website.objects.filter(
        ~Q(activity__name__exact="DECOMMISSIONED") & Q(ga4_required__name__exact="Yes")
    ).count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs
    current_datetime_now = datetime.now()
    current_datetime_now_tostring = current_datetime_now.strftime("%Y/%B")
    return render(request, 'website/websites_google_analytics.html', {
        'websites': websites,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count': table_item_count,
        'current_datetime_now_tostring': current_datetime_now_tostring,
    })


def websites_full_table(request):
    websites = Website.objects.all()
    vms = Vm.objects.all()
    table_item_count_sites = Website.objects.all().count
    table_item_count_vms = Vm.objects.all().count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs

    # ssl_date = Website.ssl_date

    return render(request, 'website/website_full_table.html', {
        'websites': websites,
        'vms': vms,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_sites': table_item_count_sites,
        'table_item_count_vms': table_item_count_vms,
    })


def websites_process_2(request):
    return render(request, 'website/websites_process_2.html')


############################# WEBSITES CRUD
def create_website(request):
    website_form = WebsiteForm(request.POST or None)
    if website_form.is_valid():
        website_form.save()
        return redirect('websites_home')
    return render(request, 'website_form.html', {'website_form': website_form})


def website_details(request, id):
    website = Website.objects.get(id=id)
    return render(request, 'website/website_details.html', {'website': website})


def update_website(request, id):
    website = Website.objects.get(id=id)
    website_form = WebsiteForm(request.POST or None, instance=website)

    if website_form.is_valid():
        website_form.save()
        return redirect('websites_home')

    return render(request, 'website_form.html', {'website_form': website_form, 'website': website})


def delete_website(request, id):
    website = Website.objects.get(id=id)
    if request.method == 'POST':
        website.delete()
        return redirect('websites_home')

    return render(request, 'website/website_delete_confirmation.html', {'website': website})

#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    SUBSITES
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################
def lddt_subsites_home(request):
    websites = Website.objects.filter(
        Q(type__name__exact="SUB-SITE")
    )
    table_item_count_subsites = Website.objects.filter(
        Q(type__name__exact="SUB-SITE")
    ).count
    myFilter = WebsiteFilter(request.GET, queryset=websites)
    shortFilter = shortwebsiteFilter(request.GET, queryset=websites)
    websites = myFilter.qs

    return render(request, 'website/subsites_home.html', {
        'websites': websites,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_subsites': table_item_count_subsites,
    })

############################# SUB-SITES CRUD

def subsite_details(request, id):
    website = Website.objects.get(id=id)
    return render(request, 'website/subsite_details.html', {'website': website})


def create_subsite(request):
    website_form = WebsiteForm(request.POST or None)
    if website_form.is_valid():
        website_form.save()
        return redirect('lddt_subsites_home')

    return render(request, 'subsite_form.html', {'website_form': website_form})


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    VM's
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def vm_home(request):
    vms = Vm.objects.all()
    table_item_count_vms = Vm.objects.count()
    myFilter = VmFilter(request.GET, queryset=vms)
    shortFilter = shortvmFilter(request.GET, queryset=vms)
    vms = myFilter.qs

    last_cron_run = cache.get("last_cron_run")

    return render(request, 'vm/vms_home.html', {
        'vms': vms,
        'myFilter': myFilter,
        'shortFilter': shortFilter,
        'table_item_count_vms': table_item_count_vms,
        'last_cron_run': last_cron_run,
    })

@require_POST
def run_vm_update(request):
    try:
        call_command('script_copy_properties')
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def vms_working_on(request):
    vms_working_on_status = Vm.objects.filter(
        Q(vm_type__name__exact="VM") & Q(vm_status__name__exact="Working On")
    )
    vms_working_on_status_count = Vm.objects.filter(
        Q(vm_type__name__exact="VM") & Q(vm_status__name__exact="Working On")
    ).count
    return render(request, 'vm/vms_working_on.html', {
        'vms_working_on_status': vms_working_on_status,
        'vms_working_on_status_count': vms_working_on_status_count,
    })


def vms_investigate(request):
    vms_investigate_status = Vm.objects.filter(
        Q(vm_type__name__exact="VM") & Q(vm_status__name__exact="Investigate")
    )
    vms_investigate_status_count = Vm.objects.filter(
        Q(vm_type__name__exact="VM") & Q(vm_status__name__exact="Investigate")
    ).count
    return render(request, 'vm/vms_investigate.html', {
        'vms_investigate_status': vms_investigate_status,
        'vms_investigate_status_count': vms_investigate_status_count,
    })


############################# VM's CRUD

def vm_details(request, id):
    vm = Vm.objects.get(id=id)
    return render(request, 'vm/vm_details.html', {'vm': vm})


def create_vm(request):
    vm_form = VmForm(request.POST or None)
    if vm_form.is_valid():
        vm_form.save()
        return redirect('vms_home')
    return render(request, 'vm_form.html', {'vm_form': vm_form})


def update_vm(request, id):
    vm = Vm.objects.get(id=id)
    vm_form = VmForm(request.POST or None, instance=vm)

    if vm_form.is_valid():
        vm_form.save()
        return redirect('vms_home')

    return render(request, 'vm_form.html', {'vm_form': vm_form, 'vm': vm})


def delete_vm(request, id):
    vm = Vm.objects.get(id=id)

    if request.method == 'POST':
        vm.delete()
        return redirect('vms_home')

    return render(request, 'vm/vm_delete_confirmation.html', {'vm': vm})


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    TUTORIALS
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def tutorials_home(request):
    return render(request, 'tutorials/tutorials_home.html')


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    WHAT CERTIFICATE
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def what_ssl_certificate(request):
    return render(request, 'tutorials/what_ssl_cert.html')


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    RENEW
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def renew_existing_certificate(request):
    return render(request, 'tutorials/renew_existing_certificate.html')


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    TECHNICAL DOCUMENTATION
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def tech_docs_home(request):
    return render(request, 'tech_docs/tech_docs_home.html')


#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    ACCESSIBILITY STATEMENT
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

def access_statement_home(request):
    access_statements = AccessStatement.objects.all()
    access_statements_all_count = AccessStatement.objects.all().count
    return render(request, 'access_statement/access_statement_home.html', {
        'access_statements': access_statements,
        'access_statements_all_count': access_statements_all_count,
    })


############################# ACCESSIBILITY STATEMENT CRUD

def create_access_statement(request):
    access_statement_form = AccessStatementForm(request.POST or None)
    if access_statement_form.is_valid():
        access_statement_form.save()
        return redirect('access_statement_home')
    return render(request, 'create_access_statement_form.html', {
        'access_statement_form': access_statement_form,
    })


def access_statement_details(request, id):
    access_statement = AccessStatement.objects.get(id=id)
    return render(request, 'access_statement/access_statement_details.html', {'access_statement': access_statement})


def update_access_statement(request, id):
    access_statement = AccessStatement.objects.get(id=id)
    access_statement_form = AccessStatementForm(request.POST or None, instance=access_statement)

    if access_statement_form.is_valid():
        access_statement_form.save()
        return redirect('access_statement_home')

    return render(request, 'create_access_statement_form.html',
                  {'access_statement_form': access_statement_form, 'access_statement': access_statement})


def delete_statement(request, id):
    access_statement = AccessStatement.objects.get(id=id)

    if request.method == 'POST':
        access_statement.delete()
        return redirect('access_statement_home')

    return render(request, 'access_statement/delete_statement.html', {'access_statement': access_statement})

#######################################################################################################################
# ---------------------------------------------------------------------------------------------------------------------
#                                                    GOOGLE ANALYTICS GA4
# ---------------------------------------------------------------------------------------------------------------------
#######################################################################################################################

# === Main view ===
def ga4_report(request):
    # ---------- LAST 12 MONTHS ----------
    months = []
    today = date.today()

    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    # ---------- GET LATEST ROW PER PROPERTY ----------
    property_ids = GoogleAnalyticsStats.objects.values_list(
        "property_id", flat=True
    ).distinct()

    properties = []
    yearly_columns = set()

    for pid in property_ids:
        latest_date = (
            GoogleAnalyticsStats.objects
            .filter(property_id=pid)
            .aggregate(Max("date"))["date__max"]
        )

        stat = GoogleAnalyticsStats.objects.get(
            property_id=pid,
            date=latest_date
        )

        # Directly use the field assuming it's already a dict or None
        monthly_data = stat.monthly_users_data or {}

        yearly_data = defaultdict(int)
        for month_key, value in monthly_data.items():
            year = int(month_key.split("-")[0])
            yearly_data[year] += int(value)
            yearly_columns.add(year)

        stat.monthly_data = monthly_data
        stat.yearly_data = yearly_data

        properties.append(stat)

    context = {
        "properties": properties,
        "months": months,
        "yearly_columns": sorted(yearly_columns),
    }

    return render(request, "ga4_reports.html", context)


def ga4_years_visits(request):
    months = []
    today = date.today()

    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    property_ids = GoogleAnalyticsStats.objects.values_list("property_id", flat=True).distinct()

    properties = []
    yearly_columns = set()

    for pid in property_ids:
        latest_date = (
            GoogleAnalyticsStats.objects
            .filter(property_id=pid)
            .aggregate(Max("date"))["date__max"]
        )
        stat = GoogleAnalyticsStats.objects.get(property_id=pid, date=latest_date)

        # Directly use the field, no json.loads needed
        monthly_data = stat.monthly_users_data or {}

        yearly_data = defaultdict(int)
        for month_key, value in monthly_data.items():
            year = int(month_key.split("-")[0])
            yearly_data[year] += int(value)
            yearly_columns.add(year)

        stat.monthly_data = monthly_data
        stat.yearly_data = yearly_data

        properties.append(stat)

    context = {
        "properties": properties,
        "months": months,
        "yearly_columns": sorted(yearly_columns),
    }

    return render(request, "ga4_pages/ga4_years_visits.html", context)

def ga4_last_12_months_sessions(request):
    months = []
    today = date.today()

    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    property_ids = GoogleAnalyticsStats.objects.values_list("property_id", flat=True).distinct()

    properties = []
    yearly_columns = set()

    for pid in property_ids:
        latest_date = (
            GoogleAnalyticsStats.objects
            .filter(property_id=pid)
            .aggregate(Max("date"))["date__max"]
        )

        stat = GoogleAnalyticsStats.objects.get(property_id=pid, date=latest_date)

        monthly_sessions = stat.monthly_sessions_data or {}

        yearly_data = defaultdict(int)
        for month_key, value in monthly_sessions.items():
            year = int(month_key.split("-")[0])
            yearly_data[year] += int(value)
            yearly_columns.add(year)

        stat.monthly_sessions = monthly_sessions
        stat.yearly_data_sessions = yearly_data

        properties.append(stat)

    context = {
        "properties": properties,
        "months": months,
        "yearly_columns": sorted(yearly_columns),
    }

    return render(request, "ga4_pages/ga4_12mts_sessions.html", context)





def _pct_change_series(values):
    """
    values: list[int]
    returns: list[float|None] same length (first None; also None if previous month is 0)
    """
    out = [None]
    for i in range(1, len(values)):
        prev = values[i - 1]
        curr = values[i]
        if prev == 0:
            out.append(None)
        else:
            out.append(round(((curr - prev) / prev) * 100.0, 2))
    return out


def top_properties_last12m_sessions_chart(request):
    today = timezone.localdate()

    # Last 12 months keys and labels
    month_keys = []
    month_labels = []
    for i in range(11, -1, -1):
        d = today - relativedelta(months=i)
        month_keys.append(d.strftime("%Y-%m"))
        month_labels.append(d.strftime("%b %Y"))

    # Latest snapshot date per property
    latest_date_subq = (
        GoogleAnalyticsStats.objects
        .filter(property_id=OuterRef("property_id"))
        .order_by("-date")
        .values("date")[:1]
    )

    # Latest row per property (so JSON reflects the most recent sync)
    latest_rows = (
        GoogleAnalyticsStats.objects
        .annotate(latest_date=Subquery(latest_date_subq))
        .filter(date=Subquery(latest_date_subq))
        .only("property_id", "property_name", "monthly_sessions_data")
    )

    ranked = []
    for row in latest_rows:
        msd = row.monthly_sessions_data or {}
        sessions_series = [int(msd.get(k, 0) or 0) for k in month_keys]
        total_12m = sum(sessions_series)  # rank by total sessions over 12 months
        ranked.append((total_12m, row.property_id, row.property_name, sessions_series))

    ranked.sort(key=lambda x: x[0], reverse=True)
    top10 = ranked[:10]

    # Chart 1: monthly sessions datasets
    sessions_datasets = []
    # Chart 2: monthly growth datasets
    growth_datasets = []

    for total_12m, pid, name, sessions_series in top10:
        sessions_datasets.append({
            "label": name,
            "data": sessions_series,
            "tension": 0.25,
        })

        growth_series = _pct_change_series(sessions_series)
        growth_datasets.append({
            "label": name,
            "data": growth_series,
            "tension": 0.25,
            "spanGaps": True,  # draw across None values
        })

    context = {
        "labels_json": json.dumps(month_labels),
        "datasets_json": json.dumps(sessions_datasets),
        "growth_datasets_json": json.dumps(growth_datasets),
    }
    return render(request, "ga4_pages/top_properties_last12m_sessions_chart.html", context)

def ga4_last_12_months_active_users(request):
    # last 12 months keys YYYY-MM
    today = date.today()
    months = []
    for i in range(11, -1, -1):
        y = today.year
        m = today.month - i
        while m <= 0:
            m += 12
            y -= 1
        months.append(f"{y}-{m:02d}")

    # latest row per property
    property_ids = GoogleAnalyticsStats.objects.values_list("property_id", flat=True).distinct()

    properties = []
    for pid in property_ids:
        latest_date = (
            GoogleAnalyticsStats.objects
            .filter(property_id=pid)
            .aggregate(Max("date"))["date__max"]
        )
        stat = GoogleAnalyticsStats.objects.get(property_id=pid, date=latest_date)

        # ✅ make it match the table template (dict with YYYY-MM keys)
        stat.monthly_data = stat.monthly_users_data or {}

        properties.append(stat)

    context = {
        "properties": sorted(properties, key=lambda s: (s.property_name or "").lower()),
        "months": months,
    }
    return render(request, "ga4_pages/ga4_12mts_active_users.html", context)







from collections import defaultdict
from dateutil.relativedelta import relativedelta

from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from lddt_app.models import GoogleAnalyticsStats


def _pct_change_series(values):
    out = [None]
    for i in range(1, len(values)):
        prev = values[i - 1]
        curr = values[i]
        if prev == 0:
            out.append(None)
        else:
            out.append(round(((curr - prev) / prev) * 100.0, 2))
    return out


def _safe_sheet_title(title: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in bad:
        title = title.replace(ch, " ")
    return title[:31] or "Sheet"


def _add_sheet(wb: Workbook, title: str, headers: list, rows: list):
    ws = wb.create_sheet(title=_safe_sheet_title(title))
    ws.append(headers)

    for r in rows:
        ws.append(r)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 40)

    return ws


def ga4_export_all_excel(request):
    today = timezone.localdate()
    month_keys = [(today - relativedelta(months=i)).strftime("%Y-%m") for i in range(11, -1, -1)]

    latest_date_subq = (
        GoogleAnalyticsStats.objects
        .filter(property_id=OuterRef("property_id"))
        .order_by("-date")
        .values("date")[:1]
    )

    latest_rows = (
        GoogleAnalyticsStats.objects
        .annotate(_latest_date=Subquery(latest_date_subq))
        .filter(date=Subquery(latest_date_subq))
        .only(
            "property_id", "property_name", "date",
            "daily_users", "monthly_users",
            "earliest_data_date",
            "monthly_users_data",
            "monthly_sessions_data",
        )
    )

    # Collect available years (from monthly_users_data)
    years_set = set()
    for r in latest_rows:
        for k in (r.monthly_users_data or {}).keys():
            try:
                years_set.add(int(k.split("-")[0]))
            except Exception:
                pass
    years = sorted(years_set)

    # ---- Sheet 1: Visits last 12 months (your existing meaning)
    headers_12m_visits = ["property_id", "property_name", "latest_date"] + month_keys + ["daily_users", "monthly_users"]
    rows_12m_visits = []
    for r in latest_rows.order_by("property_name"):
        mud = r.monthly_users_data or {}
        rows_12m_visits.append(
            [r.property_id, r.property_name, r.date.isoformat()]
            + [int(mud.get(m, 0) or 0) for m in month_keys]
            + [r.daily_users, r.monthly_users]
        )

    # ---- Sheet 2: Visits by year
    headers_year_visits = ["property_id", "property_name", "latest_date"] + [str(y) for y in years]
    rows_year_visits = []
    for r in latest_rows.order_by("property_name"):
        mud = r.monthly_users_data or {}
        yearly = defaultdict(int)
        for mk, val in mud.items():
            try:
                yearly[int(mk.split("-")[0])] += int(val or 0)
            except Exception:
                continue
        rows_year_visits.append(
            [r.property_id, r.property_name, r.date.isoformat()] + [yearly.get(y, 0) for y in years]
        )

    # ---- Sheet 3: Sessions last 12 months
    headers_12m_sessions = ["property_id", "property_name", "latest_date"] + month_keys
    rows_12m_sessions = []
    for r in latest_rows.order_by("property_name"):
        msd = r.monthly_sessions_data or {}
        rows_12m_sessions.append(
            [r.property_id, r.property_name, r.date.isoformat()]
            + [int(msd.get(m, 0) or 0) for m in month_keys]
        )

    # ✅ NEW ---- Sheet 4: Active Users last 12 months
    headers_12m_active_users = ["property_id", "property_name", "latest_date"] + month_keys
    rows_12m_active_users = []
    for r in latest_rows.order_by("property_name"):
        mud = r.monthly_users_data or {}
        rows_12m_active_users.append(
            [r.property_id, r.property_name, r.date.isoformat()]
            + [int(mud.get(m, 0) or 0) for m in month_keys]
        )

    # ---- Top 10 ranking (by total sessions)
    ranked = []
    for r in latest_rows:
        msd = r.monthly_sessions_data or {}
        series = [int(msd.get(m, 0) or 0) for m in month_keys]
        ranked.append((sum(series), r, series))

    ranked.sort(key=lambda x: x[0], reverse=True)
    top10 = ranked[:10]

    # ---- Sheet 5: Top 10 sessions
    headers_top10_sessions = ["rank", "property_id", "property_name"] + month_keys + ["total_12m"]
    rows_top10_sessions = []
    for idx, (total, r, series) in enumerate(top10, start=1):
        rows_top10_sessions.append([idx, r.property_id, r.property_name] + series + [total])

    # ---- Sheet 6: Top 10 growth %
    headers_top10_growth = ["rank", "property_id", "property_name"] + month_keys
    rows_top10_growth = []
    for idx, (total, r, series) in enumerate(top10, start=1):
        growth = _pct_change_series(series)
        growth_out = ["" if g is None else g for g in growth]
        rows_top10_growth.append([idx, r.property_id, r.property_name] + growth_out)

    # ---- Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "Visits 12 months", headers_12m_visits, rows_12m_visits)
    _add_sheet(wb, "Visits by year", headers_year_visits, rows_year_visits)
    _add_sheet(wb, "Sessions 12 months", headers_12m_sessions, rows_12m_sessions)
    _add_sheet(wb, "Active users 12m", headers_12m_active_users, rows_12m_active_users)  # ✅ NEW
    _add_sheet(wb, "Top10 Sessions 12m", headers_top10_sessions, rows_top10_sessions)
    _add_sheet(wb, "Top10 Growth %", headers_top10_growth, rows_top10_growth)

    filename = f"ga4_export_{today.isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response